from flask import Flask, render_template, request, jsonify, send_file
import os
import sys
import csv
from main import Student, Course, assign_students_to_courses, export_results, export_summary_txt
import io
from collections import Counter, defaultdict
from datetime import datetime
import pandas as pd
import random
import socket
import platform
import tempfile
import atexit
import shutil
import threading
import webbrowser
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

app = Flask(__name__)

# Globale Variable für Klassenzuteilung
class_assignment_data = {
    'students': [],        # Liste aller Schüler
    'classes': {},         # Klassen-Zuteilung: class_id -> [student_ids]
    'class_config': [],    # Klassenkonfiguration: [{id, name, language_focus, art_music}]
    'invalid_students': [] # Schüler mit unvollständigen Daten (rot markiert)
}
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max upload

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

# Betriebssystem-abhängige Konfiguration
# macOS und Linux: Nur RAM, keine persistente Speicherung
# Windows: Persistente Speicherung im data-Ordner
IS_RAM_ONLY_MODE = platform.system() in ['Darwin', 'Linux']

if IS_RAM_ONLY_MODE:
    # RAM-only Modus: Verwende temporäres Verzeichnis
    TEMP_DIR = tempfile.mkdtemp(prefix='methodentag_')
    UPLOADS_DIR = None
    RESULTS_DIR = TEMP_DIR
    
    # Cleanup beim Beenden
    def cleanup_temp_dir():
        if os.path.exists(TEMP_DIR):
            shutil.rmtree(TEMP_DIR, ignore_errors=True)
            print(f"🗑️  Temporäres Verzeichnis gelöscht: {TEMP_DIR}")
    
    atexit.register(cleanup_temp_dir)
    print(f"🔒 RAM-Only Modus aktiviert (macOS/Linux)")
    print(f"   Temporäres Verzeichnis: {TEMP_DIR}")
else:
    # Verzeichnis für persistente Datenspeicherung (nicht über Web zugänglich)
    DATA_DIR = os.path.join(SCRIPT_DIR, 'data')
    UPLOADS_DIR = os.path.join(DATA_DIR, 'uploads')
    RESULTS_DIR = os.path.join(DATA_DIR, 'results')
    
    # Erstelle Verzeichnisse falls sie nicht existieren
    os.makedirs(UPLOADS_DIR, exist_ok=True)
    os.makedirs(RESULTS_DIR, exist_ok=True)
    print("💾 Persistente Speicherung aktiviert (Windows)")

# Globale Variable für aktuelle Daten (im RAM für schnellen Zugriff)
current_data = {
    'students': None,
    'courses': None,
    'filename': None,
    'upload_path': None  # Pfad zur gespeicherten Upload-Datei
}

# Timer für automatisches Löschen der Dateien (nur im RAM-Only Modus)
AUTO_DELETE_MINUTES = 10
delete_timer = None
delete_timestamp = None

def delete_temp_files():
    """Löscht die temporären Download-Dateien nach Ablauf der Zeit"""
    global delete_timestamp
    if IS_RAM_ONLY_MODE and RESULTS_DIR and os.path.exists(RESULTS_DIR):
        print(f"⏰ Auto-Delete Timer abgelaufen - lösche temporäre Dateien")
        for filename in ['zuteilung_schueler.csv', 'zuteilung_kurse.csv', 'zusammenfassung.txt']:
            filepath = os.path.join(RESULTS_DIR, filename)
            if os.path.exists(filepath):
                os.remove(filepath)
                print(f"   🗑️  Gelöscht: {filename}")
        delete_timestamp = None

def start_delete_timer():
    """Startet den Timer für automatisches Löschen"""
    global delete_timer, delete_timestamp
    
    # Stoppe vorherigen Timer falls vorhanden
    if delete_timer is not None:
        delete_timer.cancel()
    
    if IS_RAM_ONLY_MODE:
        delete_timestamp = datetime.now().timestamp() + (AUTO_DELETE_MINUTES * 60)
        delete_timer = threading.Timer(AUTO_DELETE_MINUTES * 60, delete_temp_files)
        delete_timer.daemon = True
        delete_timer.start()
        print(f"⏰ Auto-Delete Timer gestartet: Dateien werden in {AUTO_DELETE_MINUTES} Minuten gelöscht")

def load_data_from_csv_content(file_content):
    """Lädt CSV-Daten direkt aus dem Speicher"""
    students = []
    all_courses = set()
    
    # Dekodiere Bytes zu String falls nötig
    if isinstance(file_content, bytes):
        file_content = file_content.decode('utf-8')
    
    # Verwende StringIO für In-Memory CSV-Parsing
    csv_file = io.StringIO(file_content)
    reader = csv.reader(csv_file, delimiter=';')
    
    # Überspringe die Kopfzeile
    next(reader, None)
    
    for row in reader:
        # Stelle sicher, dass die Zeile genug Spalten hat
        if len(row) < 7:
            continue
        
        # Spalten: 0=Nachname, 1=Vorname, 2=Klasse, 3-6=Wünsche 1-4
        lastname = row[0].strip()
        firstname = row[1].strip()
        klasse = row[2].strip()
        
        # Ignoriere Zeilen mit "unbekannt" oder "unknown" als Klasse (case-insensitive)
        if klasse.lower() in ['unbekannt', 'unknown']:
            continue
        
        # Sammle die Wünsche aus Spalten 3-6
        wishes = []
        for i in range(3, 7):
            if row[i] and row[i].strip():
                wish = row[i].strip()
                wishes.append(wish)
                all_courses.add(wish)
        
        student = Student(lastname, firstname, klasse, wishes)
        students.append(student)
    
    # Erstelle Course-Objekte
    courses = {name: Course(name) for name in all_courses}
    
    return students, courses

def load_data_from_file(filepath):
    """Lädt die CSV-Datei von Disk (Fallback für daten.csv)"""
    with open(filepath, 'r', encoding='utf-8') as f:
        content = f.read()
    return load_data_from_csv_content(content)

@app.route('/')
def index():
    return render_template('index.html', ram_only_mode=IS_RAM_ONLY_MODE, os_name=platform.system())

@app.route('/api/upload', methods=['POST'])
def upload_file():
    """Lädt eine CSV-Datei hoch und speichert sie persistent im data/uploads Verzeichnis"""
    print("=" * 60)
    print("📤 UPLOAD REQUEST empfangen")
    print(f"Request files: {request.files}")
    print(f"Request form: {request.form}")
    print(f"Request method: {request.method}")
    print(f"Content-Type: {request.content_type}")
    
    if 'file' not in request.files:
        error_msg = 'Keine Datei im Request gefunden (key "file" fehlt)'
        print(f"❌ ERROR: {error_msg}")
        print(f"Verfügbare Keys: {list(request.files.keys())}")
        return jsonify({
            'success': False,
            'error': error_msg
        }), 400
    
    file = request.files['file']
    print(f"📁 Datei gefunden: {file.filename}")
    
    if file.filename == '':
        error_msg = 'Dateiname ist leer'
        print(f"❌ ERROR: {error_msg}")
        return jsonify({
            'success': False,
            'error': 'Keine Datei ausgewählt'
        }), 400
    
    # Prüfe Dateiendung
    if not file.filename.lower().endswith('.csv'):
        error_msg = f'Ungültiger Dateityp: {file.filename}'
        print(f"❌ ERROR: {error_msg}")
        return jsonify({
            'success': False,
            'error': 'Ungültiger Dateityp. Nur CSV-Dateien sind erlaubt.'
        }), 400
    
    try:
        # Lese Datei direkt in den Speicher
        print("📖 Lese Dateiinhalt...")
        file_content = file.read()
        print(f"✓ Datei gelesen: {len(file_content)} Bytes")
        
        # Speichere Datei nur auf Windows persistent
        upload_path = None
        if not IS_RAM_ONLY_MODE:
            # Erstelle Zeitstempel für eindeutigen Dateinamen
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            safe_filename = f"{timestamp}_{file.filename}"
            upload_path = os.path.join(UPLOADS_DIR, safe_filename)
            
            # Speichere Datei persistent
            print(f"💾 Speichere Datei nach: {upload_path}")
            with open(upload_path, 'wb') as f:
                f.write(file_content)
            print(f"✓ Datei gespeichert")
        else:
            print(f"🔒 RAM-Only Modus: Datei wird NICHT gespeichert")
        
        # Versuche die Daten zu laden und zu validieren
        print("🔄 Parse CSV-Daten...")
        students, courses = load_data_from_csv_content(file_content)
        print(f"✓ CSV geparst: {len(students)} Schüler, {len(courses)} Kurse")
        
        # Speichere in globaler Variable für schnellen Zugriff
        current_data['students'] = students
        current_data['courses'] = courses
        current_data['filename'] = file.filename
        current_data['upload_path'] = upload_path
        
        print("✅ Upload erfolgreich!")
        print("=" * 60)
        
        return jsonify({
            'success': True,
            'filename': file.filename,
            'studentCount': len(students),
            'courseCount': len(courses)
        })
    except Exception as e:
        error_msg = f'Fehler beim Laden der Datei: {str(e)}'
        print(f"❌ EXCEPTION: {error_msg}")
        print(f"Exception type: {type(e).__name__}")
        import traceback
        print(traceback.format_exc())
        print("=" * 60)
        return jsonify({
            'success': False,
            'error': error_msg
        }), 400

@app.route('/api/analyze', methods=['POST'])
def analyze():
    """Analysiert die Daten ohne Zuteilung durchzuführen"""
    try:
        # Prüfe ob Daten hochgeladen wurden
        if current_data['students'] is None or current_data['courses'] is None:
            return jsonify({
                'success': False,
                'error': 'Bitte laden Sie zuerst eine CSV-Datei hoch.'
            }), 400
        
        students = current_data['students']
        courses = current_data['courses']
        
        # Sammle Kursstatistiken
        course_list = []
        for name in sorted(courses.keys()):
            # Zähle wie oft dieser Kurs gewünscht wurde
            wish_count = 0
            wish_priorities = {1: 0, 2: 0, 3: 0, 4: 0}
            for student in students:
                for i, wish in enumerate(student.wishes, 1):
                    if wish == name:
                        wish_count += 1
                        wish_priorities[i] += 1
            
            course_list.append({
                'name': name,
                'wishCount': wish_count,
                'priorities': wish_priorities
            })
        
        return jsonify({
            'success': True,
            'studentCount': len(students),
            'courseCount': len(courses),
            'courses': course_list
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/assign', methods=['POST'])
def assign():
    """Führt die Zuteilung durch"""
    try:
        data = request.get_json()
        max_students = data.get('maxStudents')
        equal_distribution = data.get('equalDistribution', False)
        course_limits = data.get('courseLimits', {})  # Individuelle Limits pro Kurs
        
        # Konvertiere max_students (globales Limit)
        if max_students is not None and max_students != '':
            max_students = int(max_students)
        else:
            max_students = None
        
        # Prüfe ob Daten hochgeladen wurden
        if current_data['students'] is None or current_data['courses'] is None:
            return jsonify({
                'success': False,
                'error': 'Bitte laden Sie zuerst eine CSV-Datei hoch.'
            }), 400
        
        # Erstelle neue Course-Objekte für die Zuteilung (Deep Copy)
        students = current_data['students']
        courses = {name: Course(name) for name in current_data['courses'].keys()}
        
        # Setze alle Zuweisungen zurück
        for student in students:
            student.assigned_courses = []
            student.fulfilled_wish_numbers = []
        
        fulfilled_wishes, unfulfilled_students, total_wishes = assign_students_to_courses(
            students, courses, max_students, equal_distribution, course_limits
        )
        
        # Exportiere Ergebnisse
        if not IS_RAM_ONLY_MODE:
            # Windows: Speichere persistent in results Verzeichnis + Kopie für Downloads
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            result_subdir = os.path.join(RESULTS_DIR, timestamp)
            os.makedirs(result_subdir, exist_ok=True)
            
            print(f"💾 Speichere Ergebnisse nach: {result_subdir}")
            export_results(students, courses, output_dir=result_subdir)
            export_summary_txt(students, courses, fulfilled_wishes, unfulfilled_students, total_wishes, max_students, output_dir=result_subdir)
            
            # Kopiere die Dateien auch ins Hauptverzeichnis für die Download-API
            for filename in ['zuteilung_schueler.csv', 'zuteilung_kurse.csv', 'zusammenfassung.txt']:
                src = os.path.join(result_subdir, filename)
                dst = os.path.join(SCRIPT_DIR, filename)
                shutil.copy2(src, dst)
        else:
            # macOS/Linux: Nur temporär im temp-Verzeichnis speichern
            print(f"🔒 RAM-Only Modus: Temporäre Dateien im RAM-Verzeichnis")
            export_results(students, courses, output_dir=RESULTS_DIR)
            export_summary_txt(students, courses, fulfilled_wishes, unfulfilled_students, total_wishes, max_students, output_dir=RESULTS_DIR)
            
            # Starte Auto-Delete Timer
            start_delete_timer()
        
        # Berechne Statistiken
        total_fulfilled = sum(fulfilled_wishes.values())
        fulfillment_rate = (total_fulfilled / total_wishes * 100) if total_wishes > 0 else 0
        
        wishes_per_student = Counter()
        for student in students:
            num_fulfilled = len(student.fulfilled_wish_numbers)
            wishes_per_student[num_fulfilled] += 1
        
        students_without_wishes = [s for s in students if not s.fulfilled_wish_numbers]
        
        # Kursgrößen
        course_stats = []
        for course_name, course in sorted(courses.items()):
            total = course.get_total_students()
            t0 = course.get_students_in_timeslot(0)
            t1 = course.get_students_in_timeslot(1)
            t2 = course.get_students_in_timeslot(2)
            course_stats.append({
                'name': course_name,
                'total': total,
                'timeslot1': t0,
                'timeslot2': t1,
                'timeslot3': t2
            })
        
        course_stats.sort(key=lambda x: x['total'], reverse=True)
        avg_students = sum(c['total'] for c in course_stats) / len(course_stats) if course_stats else 0
        
        return jsonify({
            'success': True,
            'stats': {
                'studentCount': len(students),
                'courseCount': len(courses),
                'totalWishes': total_wishes,
                'fulfilledWishes': total_fulfilled,
                'fulfillmentRate': round(fulfillment_rate, 1),
                'wishPriorities': {str(k): v for k, v in fulfilled_wishes.items()},
                'wishesPerStudent': {str(k): v for k, v in wishes_per_student.items()},
                'studentsWithoutWishes': len(students_without_wishes),
                'averageCourseSize': round(avg_students, 1),
                'courseStats': course_stats  # Alle Kurse anzeigen
            }
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/download/<filename>')
def download(filename):
    """Download generierter Dateien (CSV oder Excel)"""
    allowed_files = [
        'zuteilung_schueler.csv', 'zuteilung_kurse.csv', 
        'zuteilung_schueler.xlsx', 'zuteilung_kurse.xlsx',
        'zusammenfassung.txt'
    ]
    
    if filename not in allowed_files:
        return jsonify({'error': 'File not found'}), 404
    
    # Bestimme den richtigen Pfad basierend auf dem Modus
    download_dir = RESULTS_DIR if IS_RAM_ONLY_MODE else SCRIPT_DIR
    
    # Wenn Excel angefragt wird, konvertiere CSV zu Excel
    if filename.endswith('.xlsx'):
        csv_filename = filename.replace('.xlsx', '.csv')
        csv_path = os.path.join(download_dir, csv_filename)
        
        if not os.path.exists(csv_path):
            return jsonify({'error': 'File not generated yet'}), 404
        
        try:
            # Konvertiere CSV zu Excel
            df = pd.read_csv(csv_path, encoding='utf-8')
            
            # Erstelle Excel-Datei im Speicher
            excel_buffer = io.BytesIO()
            with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Zuteilung')
                
                # Auto-size columns für bessere Lesbarkeit
                worksheet = writer.sheets['Zuteilung']
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)  # Max 50 Zeichen
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            excel_buffer.seek(0)
            
            return send_file(
                excel_buffer,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=filename
            )
        except Exception as e:
            return jsonify({'error': f'Excel conversion failed: {str(e)}'}), 500
    
    # Standard CSV/TXT Download
    file_path = os.path.join(download_dir, filename)
    
    if not os.path.exists(file_path):
        return jsonify({'error': 'File not generated yet'}), 404
    
    return send_file(file_path, as_attachment=True)

@app.route('/api/clear', methods=['POST'])
def clear_data():
    """Löscht die temporären Daten aus dem RAM (nicht die persistenten Dateien)"""
    global delete_timer, delete_timestamp
    
    current_data['students'] = None
    current_data['courses'] = None
    current_data['filename'] = None
    current_data['upload_path'] = None
    
    # Stoppe Timer falls aktiv
    if delete_timer is not None:
        delete_timer.cancel()
        delete_timer = None
        delete_timestamp = None
    
    return jsonify({
        'success': True,
        'message': 'Daten wurden aus dem Speicher gelöscht (persistente Dateien bleiben erhalten)'
    })

@app.route('/api/delete-timestamp', methods=['GET'])
def get_delete_timestamp():
    """Gibt den Zeitstempel zurück, wann die Dateien gelöscht werden"""
    return jsonify({
        'deleteTimestamp': delete_timestamp,
        'autoDeleteMinutes': AUTO_DELETE_MINUTES,
        'ramOnlyMode': IS_RAM_ONLY_MODE
    })

# ============================================================================
# KLASSENZUTEILUNG API
# ============================================================================

def _normalize_name(name):
    """Normalisiert einen Namen für den Vergleich (lowercase, getrimmt)"""
    return ' '.join(str(name).strip().lower().split())


def parse_wunschpartner(raw):
    """Zerlegt eine Wunschpartner-Angabe in einzelne Namen.

    Mehrere Wunschpartner werden durch Komma, Semikolon, Schrägstrich,
    Zeilenumbruch oder das Wort "und" getrennt.
    """
    if not raw:
        return []
    text = str(raw)
    # Vereinheitliche Trennzeichen zu Komma
    for sep in ['\n', '\r', ';', '/', '|', ' und ', ' & ', '&']:
        text = text.replace(sep, ',')
    parts = [p.strip() for p in text.split(',')]
    return [p for p in parts if p]


def resolve_wunschpartner(students):
    """Ordnet die Wunschpartner-Namen jedes Schülers den passenden Schüler-IDs zu.

    Setzt für jeden Schüler 'wunsch_names' (geparste Namen) und 'wunsch_ids'
    (gefundene IDs der Wunschpartner). Die Zuordnung ist wechselseitig nicht
    erzwungen – nur tatsächlich gefundene Namen werden verlinkt.
    """
    # Index zum schnellen Nachschlagen anlegen
    by_full = {}          # "vorname nachname" -> [ids]
    by_full_rev = {}      # "nachname vorname" -> [ids]
    by_first = defaultdict(list)  # "vorname" -> [ids]
    for s in students:
        full = _normalize_name(f"{s['vorname']} {s['nachname']}")
        rev = _normalize_name(f"{s['nachname']} {s['vorname']}")
        by_full.setdefault(full, []).append(s['id'])
        by_full_rev.setdefault(rev, []).append(s['id'])
        if s['vorname']:
            by_first[_normalize_name(s['vorname'])].append(s['id'])

    for s in students:
        names = parse_wunschpartner(s.get('wunschpartner_raw', ''))
        s['wunsch_names'] = names
        found_ids = []
        for name in names:
            key = _normalize_name(name)
            ids = by_full.get(key) or by_full_rev.get(key)
            if not ids:
                # Fallback: eindeutiger Vorname
                first_matches = by_first.get(key, [])
                if len(first_matches) == 1:
                    ids = first_matches
            if ids:
                for sid in ids:
                    if sid != s['id'] and sid not in found_ids:
                        found_ids.append(sid)
        s['wunsch_ids'] = found_ids


@app.route('/api/klassenzuteilung/upload', methods=['POST'])
def upload_class_list():
    """Lädt eine Excel-Datei mit Schülerliste für Klassenzuteilung hoch"""
    global class_assignment_data
    
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'Keine Datei hochgeladen'}), 400
        
        file = request.files['file']
        
        if file.filename == '':
            return jsonify({'success': False, 'error': 'Keine Datei ausgewählt'}), 400
        
        if not file.filename.endswith('.xlsx'):
            return jsonify({'success': False, 'error': 'Nur .xlsx Dateien werden unterstützt'}), 400
        
        # Load the workbook
        wb = load_workbook(file, data_only=False)
        ws = wb.active
        
        print(f"\n{'='*60}")
        print(f"📋 KLASSENZUTEILUNG UPLOAD: {file.filename}")
        print(f"Aktives Sheet: '{ws.title}'")
        print(f"Alle Sheets: {wb.sheetnames}")
        print(f"Dimensionen: {ws.dimensions}")
        print(f"Max Row: {ws.max_row}, Max Col: {ws.max_column}")
        
        # Finde alle relevanten Spalten in der ersten Zeile
        column_mapping = {}
        required_columns = ['nachnamen', 'vornamen', 'geschlecht', 'extern / intern ', '2. FS', 'ku/mu']
        optional_columns = ['bisherige klasse']
        
        first_row = list(ws[1])
        print(f"\n📊 Erste Zeile ({len(first_row)} Spalten):")
        for col_idx, cell in enumerate(first_row, start=1):
            print(f"  Spalte {col_idx}: '{cell.value}' (type: {type(cell.value).__name__})")
        
        for col_idx, cell in enumerate(first_row, start=1):
            if cell.value:
                cell_value = str(cell.value).strip().lower()
                # Exakte Matches oder enthält
                if 'nachnamen' in cell_value or cell_value == 'nachname':
                    column_mapping['nachname'] = col_idx
                elif 'vornamen' in cell_value or cell_value == 'vorname':
                    column_mapping['vorname'] = col_idx
                elif 'geschlecht' in cell_value:
                    column_mapping['geschlecht'] = col_idx
                elif 'extern' in cell_value and 'intern' in cell_value:
                    # Matches "extern / intern", "intern / extern", "extern/intern", etc.
                    column_mapping['intern_extern'] = col_idx
                elif '2. fs' in cell_value or '2.fs' in cell_value or '2. fa' in cell_value or '2.fa' in cell_value or 'fremdsprache' in cell_value:
                    column_mapping['fremdsprache'] = col_idx
                elif 'ku/mu' in cell_value or cell_value == 'kumu' or cell_value == 'ku / mu':
                    column_mapping['kunst_musik'] = col_idx
<<<<<<< HEAD
                elif 'aufnahme' in cell_value:
                    column_mapping['aufnahme'] = col_idx
                elif 'wunschpartner' in cell_value or 'wunsch' in cell_value:
                    column_mapping['wunschpartner'] = col_idx
=======
                elif 'reli' in cell_value and 'ethik' in cell_value:
                    column_mapping['reli_ethik'] = col_idx
                elif cell_value == 'zf' or cell_value == 'zusatzfach':
                    column_mapping['zusatzfach'] = col_idx
>>>>>>> 12d6a2ed4dd06facda55ba5696c9a5b42f786793
                elif 'bisherige klasse' in cell_value or ('klasse' in cell_value and 'bisherige' in cell_value):
                    column_mapping['bisherige_klasse'] = col_idx
        
        # Prüfe ob alle erforderlichen Spalten vorhanden sind
        missing = []
        if 'nachname' not in column_mapping:
            missing.append('Nachnamen')
        if 'vorname' not in column_mapping:
            missing.append('Vornamen')
        if 'geschlecht' not in column_mapping:
            missing.append('Geschlecht')
        if 'intern_extern' not in column_mapping:
            missing.append('extern / intern')
        if 'fremdsprache' not in column_mapping:
            missing.append('2. FS')
        if 'kunst_musik' not in column_mapping:
            missing.append('KU/MU')
        if 'aufnahme' not in column_mapping:
            missing.append('Aufnahme')
        
        print(f"\n🗺️  Column Mapping: {column_mapping}")
        
        if missing:
            print(f"❌ Fehlende Spalten: {missing}")
            return jsonify({
                'success': False, 
                'error': f'Fehlende Spalten in Zeile 1: {", ".join(missing)}'
            }), 400
        
        students = []
        invalid_students = []
        student_id = 0
        skipped_red = 0
        skipped_empty = 0
        total_rows_checked = 0
        
        for row_num, row in enumerate(ws.iter_rows(min_row=2), start=2):
            total_rows_checked += 1
            # Prüfe ob Zeile rot markiert ist
            is_red = False
            for cell in row:
                if cell.fill and cell.fill.start_color:
                    color = cell.fill.start_color
                    if hasattr(color, 'rgb') and color.rgb:
                        rgb_str = str(color.rgb)
                        if len(rgb_str) == 8:
                            rgb_str = rgb_str[2:]
                        try:
                            r = int(rgb_str[0:2], 16)
                            g = int(rgb_str[2:4], 16)
                            b = int(rgb_str[4:6], 16)
                            if r > 180 and r > (g + 50) and r > (b + 50):
                                is_red = True
                                break
                        except (ValueError, IndexError):
                            pass
            
            if is_red:
                print(f"  🔴 Zeile {row_num}: Übersprungen (rot markiert)")
                skipped_red += 1
                continue  # Überspringe rot markierte Zeilen
            
            # Extrahiere Daten
            def get_cell_value(col_name):
                if col_name in column_mapping:
                    idx = column_mapping[col_name] - 1
                    if idx < len(row) and row[idx].value:
                        return str(row[idx].value).strip()
                return ''
            
            nachname = get_cell_value('nachname')
            vorname = get_cell_value('vorname')
            geschlecht = get_cell_value('geschlecht').lower()
            intern_extern = get_cell_value('intern_extern').lower()
            fremdsprache = get_cell_value('fremdsprache')
            kunst_musik = get_cell_value('kunst_musik')
            reli_ethik = get_cell_value('reli_ethik')
            zusatzfach = get_cell_value('zusatzfach')
            bisherige_klasse = get_cell_value('bisherige_klasse')
            aufnahme = get_cell_value('aufnahme')
            wunschpartner_raw = get_cell_value('wunschpartner')
            
            # Überspringe leere Zeilen
            if not nachname and not vorname:
                skipped_empty += 1
                if total_rows_checked <= 5:  # Nur erste 5 leere Zeilen loggen
                    print(f"  ⬜ Zeile {row_num}: Übersprungen (Name leer)")
                continue
            
<<<<<<< HEAD
            # Filter: Nur Schüler mit "Aufnahme" in der Aufnahme-Spalte werden eingebracht
            if 'aufnahme' not in aufnahme.strip().lower():
                continue
=======
            # Log erste paar Schüler
            if student_id < 3:
                print(f"  👤 Zeile {row_num}: {vorname} {nachname}, Geschl={geschlecht}, I/E={intern_extern}, FS={fremdsprache}, KuMu={kunst_musik}, Reli={reli_ethik}, ZF={zusatzfach}")
>>>>>>> 12d6a2ed4dd06facda55ba5696c9a5b42f786793
            
            # Normalisiere Werte
            if geschlecht in ['m', 'männlich', 'male']:
                geschlecht = 'm'
            elif geschlecht in ['w', 'weiblich', 'female', 'f']:
                geschlecht = 'w'
            else:
                geschlecht = ''
            
            # Normalisiere intern/extern
            if 'intern' in intern_extern and 'extern' not in intern_extern:
                intern_extern_normalized = 'intern'
            elif 'extern' in intern_extern:
                intern_extern_normalized = 'extern'
            else:
                intern_extern_normalized = ''
            
            # Normalisiere Fremdsprache
            fremdsprache_upper = fremdsprache.upper() if fremdsprache else ''
            if fremdsprache_upper in ['F', 'FRANZÖSISCH', 'FRANZ']:
                fremdsprache = 'F'
            elif fremdsprache_upper in ['L', 'LATEIN']:
                fremdsprache = 'L'
            elif fremdsprache_upper in ['L0', 'LATEIN0']:
                fremdsprache = 'L0'
            elif fremdsprache_upper in ['SPA', 'SPANISCH']:
                fremdsprache = 'Spa'
            else:
                fremdsprache = fremdsprache if fremdsprache else ''
            
            # Normalisiere Kunst/Musik
            kunst_musik_lower = kunst_musik.lower() if kunst_musik else ''
            if kunst_musik_lower in ['ku', 'kunst']:
                kunst_musik = 'Ku'
            elif kunst_musik_lower in ['mu', 'musik']:
                kunst_musik = 'Mu'
            else:
                kunst_musik = kunst_musik if kunst_musik else ''
            
            # Normalisiere Reli/Ethik
            reli_ethik_lower = reli_ethik.lower() if reli_ethik else ''
            if reli_ethik_lower in ['rev', 'evangelisch', 'ev']:
                reli_ethik = 'Rev'
            elif reli_ethik_lower in ['rka', 'katholisch', 'ka']:
                reli_ethik = 'Rka'
            elif reli_ethik_lower in ['eth', 'ethik']:
                reli_ethik = 'Eth'
            else:
                reli_ethik = reli_ethik if reli_ethik else ''
            
            # Normalisiere Zusatzfach
            zusatzfach_lower = zusatzfach.lower() if zusatzfach else ''
            if zusatzfach_lower in ['ek', 'erdkunde']:
                zusatzfach = 'Ek'
            elif zusatzfach_lower in ['phil', 'philosophie']:
                zusatzfach = 'Phil'
            elif zusatzfach_lower in ['info', 'informatik']:
                zusatzfach = 'Info'
            elif zusatzfach_lower in ['spa', 'spanisch']:
                zusatzfach = 'Spa'
            elif zusatzfach_lower in ['bili', 'bilingual']:
                zusatzfach = 'Bili'
            else:
                zusatzfach = zusatzfach if zusatzfach else ''
            
            student = {
                'id': student_id,
                'row': row_num,
                'nachname': nachname,
                'vorname': vorname,
                'geschlecht': geschlecht,
                'intern_extern': intern_extern_normalized,
                'fremdsprache': fremdsprache,
                'kunst_musik': kunst_musik,
                'reli_ethik': reli_ethik,
                'zusatzfach': zusatzfach,
                'bisherige_klasse': bisherige_klasse,
                'wunschpartner_raw': wunschpartner_raw,
                'wunsch_names': [],
                'wunsch_ids': [],
                'assigned_class': None
            }
            
            # Prüfe ob alle erforderlichen Felder vorhanden sind
            is_valid = True
            missing_fields = []
            
            if not nachname:
                missing_fields.append('Nachname')
                is_valid = False
            if not vorname:
                missing_fields.append('Vorname')
                is_valid = False
            if not geschlecht:
                missing_fields.append('Geschlecht')
                is_valid = False
            if not intern_extern_normalized:
                missing_fields.append('intern/extern')
                is_valid = False
            if not kunst_musik:
                missing_fields.append('KU/MU')
                is_valid = False
            
            if is_valid:
                students.append(student)
            else:
                student['missing_fields'] = missing_fields
                invalid_students.append(student)
                print(f"  ⚠️  Zeile {row_num}: UNGÜLTIG - {vorname} {nachname} - fehlende Felder: {missing_fields}")
            
            student_id += 1
        
<<<<<<< HEAD
        # Wunschpartner-Namen den Schüler-IDs zuordnen (nur gültige Schüler)
        resolve_wunschpartner(students)
=======
        print(f"\n📈 Zusammenfassung:")
        print(f"  Zeilen geprüft: {total_rows_checked}")
        print(f"  Rot übersprungen: {skipped_red}")
        print(f"  Leer übersprungen: {skipped_empty}")
        print(f"  Gültige Schüler: {len(students)}")
        print(f"  Ungültige Schüler: {len(invalid_students)}")
        print(f"{'='*60}")
>>>>>>> 12d6a2ed4dd06facda55ba5696c9a5b42f786793
        
        # Statistiken berechnen
        male_count = sum(1 for s in students if s['geschlecht'] == 'm')
        female_count = sum(1 for s in students if s['geschlecht'] == 'w')
        intern_count = sum(1 for s in students if s['intern_extern'] == 'intern')
        extern_count = sum(1 for s in students if s['intern_extern'] == 'extern')
        
        # Fremdsprachen-Statistik
        language_stats = defaultdict(int)
        for s in students:
            if s['fremdsprache']:
                language_stats[s['fremdsprache']] += 1
        
        # Kunst/Musik-Statistik
        km_stats = defaultdict(int)
        for s in students:
            if s['kunst_musik']:
                km_stats[s['kunst_musik']] += 1
        
        # Reli/Ethik-Statistik
        reli_stats = defaultdict(int)
        for s in students:
            if s['reli_ethik']:
                reli_stats[s['reli_ethik']] += 1
        
        # Zusatzfach-Statistik
        zf_stats = defaultdict(int)
        for s in students:
            if s['zusatzfach']:
                zf_stats[s['zusatzfach']] += 1
        
        # Speichere Daten
        class_assignment_data['students'] = students
        class_assignment_data['invalid_students'] = invalid_students
        class_assignment_data['classes'] = {}
        class_assignment_data['class_config'] = []
        
        return jsonify({
            'success': True,
            'filename': file.filename,
            'studentCount': len(students),
            'invalidCount': len(invalid_students),
            'invalidStudents': invalid_students,
            'stats': {
                'male': male_count,
                'female': female_count,
                'intern': intern_count,
                'extern': extern_count,
                'languages': dict(language_stats),
                'kunstMusik': dict(km_stats),
                'reliEthik': dict(reli_stats),
                'zusatzfach': dict(zf_stats)
            }
        })
    
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/klassenzuteilung/configure', methods=['POST'])
def configure_classes():
    """Konfiguriert die Klassen (Anzahl, Sprachfokus, Kunst/Musik)"""
    global class_assignment_data
    
    try:
        data = request.get_json()
        class_configs = data.get('classes', [])
        
        if not class_configs:
            return jsonify({'success': False, 'error': 'Keine Klassenkonfiguration angegeben'}), 400
        
        # Validiere Konfiguration
        for i, config in enumerate(class_configs):
            if 'name' not in config:
                config['name'] = f'Klasse {i + 1}'
            if 'language_focus' not in config:
                config['language_focus'] = ''
            if 'art_music' not in config:
                config['art_music'] = ''
            if 'reli_focus' not in config:
                config['reli_focus'] = ''
            if 'zf_focus' not in config:
                config['zf_focus'] = ''
            config['id'] = i
        
        class_assignment_data['class_config'] = class_configs
        class_assignment_data['classes'] = {i: [] for i in range(len(class_configs))}
        
        return jsonify({
            'success': True,
            'classCount': len(class_configs),
            'classes': class_configs
        })
    
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/klassenzuteilung/assign', methods=['POST'])
def auto_assign_classes():
    """Führt die automatische Klassenzuteilung durch"""
    global class_assignment_data
    
    try:
        students = class_assignment_data['students']
        class_config = class_assignment_data['class_config']
        
        if not students:
            return jsonify({'success': False, 'error': 'Keine Schülerdaten vorhanden'}), 400
        
        if not class_config:
            return jsonify({'success': False, 'error': 'Keine Klassenkonfiguration vorhanden'}), 400
        
        num_classes = len(class_config)
        
        # Reset alle Zuweisungen
        for s in students:
            s['assigned_class'] = None
        
        classes = {i: [] for i in range(num_classes)}
        
        # Schritt 1: Mische Schüler zufällig (damit keine Reihenfolge-Vorteile entstehen)
        shuffled_students = students.copy()
        random.shuffle(shuffled_students)
        
        # Berechne Zielgröße pro Klasse
        target_size = len(students) // num_classes
        remainder = len(students) % num_classes  # Einige Klassen bekommen 1 mehr
        
        # Erstelle Liste der max. Größe pro Klasse
        max_class_sizes = {}
        for i in range(num_classes):
            # Erste 'remainder' Klassen bekommen einen Schüler mehr
            max_class_sizes[i] = target_size + (1 if i < remainder else 0)
        
        # Prüfe welche Klassen Kunst/Musik-Fokus haben
        ku_classes = [i for i, c in enumerate(class_config) if c.get('art_music') == 'Ku']
        mu_classes = [i for i, c in enumerate(class_config) if c.get('art_music') == 'Mu']
        mixed_classes = [i for i, c in enumerate(class_config) if not c.get('art_music')]
        
        # Gruppiere Schüler nach Kunst/Musik
        ku_students = [s for s in shuffled_students if s.get('kunst_musik') == 'Ku']
        mu_students = [s for s in shuffled_students if s.get('kunst_musik') == 'Mu']
        other_students = [s for s in shuffled_students if s.get('kunst_musik') not in ('Ku', 'Mu')]
        
        random.shuffle(ku_students)
        random.shuffle(mu_students)
        random.shuffle(other_students)
        
        # Hilfsfunktion zur Berechnung des Scores für Zuweisung (ohne Kunst/Musik-Check)
        def assignment_score(student, class_id, current_classes):
            config = class_config[class_id]
            class_students = current_classes[class_id]
            
            score = 100  # Basiswert
            
            # Sprachfokus:
            # L = nur Latein (Fortgeschritten)
            # L0 = nur Latein 0 (Neuanfänger)
            # L_all = Latein + Latein 0 (beide)
            # F = Französisch
            # Spa = Spanisch
            lang_focus = config.get('language_focus', '')
            student_lang = student['fremdsprache']
            
            if lang_focus:
                if lang_focus == 'L' and student_lang == 'L':
                    score += 30
                elif lang_focus == 'L0' and student_lang == 'L0':
                    score += 30
                elif lang_focus == 'L_all' and student_lang in ['L', 'L0']:
                    score += 30
                elif lang_focus == 'F' and student_lang == 'F':
                    score += 30
                elif lang_focus == 'Spa' and student_lang == 'Spa':
                    score += 30
                else:
                    # Schüler passt nicht zum Sprachfokus
                    score -= 20
            
            # Kunst/Musik - Fokus beachten, sonst gleichmäßig verteilen
            km_focus = config.get('art_music', '')
            student_km = student.get('kunst_musik', '')
            
            if km_focus:
                # Klasse hat Ku/Mu-Fokus
                if student_km == km_focus:
                    score += 25  # Bonus: Schüler passt zum Fokus
                else:
                    score -= 15  # Strafe: Schüler passt nicht
            elif class_students and student_km:
                # Kein Fokus (Gemischt) -> gleichmäßig verteilen
                km_in_class = sum(1 for s in class_students if s.get('kunst_musik') == student_km)
                total_km = sum(1 for s in students if s.get('kunst_musik') == student_km)
                global_km_ratio = total_km / len(students) if students else 0
                
                new_ratio = (km_in_class + 1) / (len(class_students) + 1)
                km_penalty = abs(global_km_ratio - new_ratio) * 30
                score -= km_penalty
            
            # Geschlechterbalance
            if class_students:
                male_in_class = sum(1 for s in class_students if s['geschlecht'] == 'm')
                total_in_class = len(class_students)
                current_male_ratio = male_in_class / total_in_class
                
                # Berechne globale m/w Quote
                total_male = sum(1 for s in students if s['geschlecht'] == 'm')
                global_male_ratio = total_male / len(students) if students else 0.5
                
                if student['geschlecht'] == 'm':
                    new_ratio = (male_in_class + 1) / (total_in_class + 1)
                else:
                    new_ratio = male_in_class / (total_in_class + 1)
                
                # Strafe für Abweichung von globaler Quote
                balance_penalty = abs(global_male_ratio - new_ratio) * 50
                score -= balance_penalty
            
            # Intern/Extern Balance
            if class_students:
                extern_in_class = sum(1 for s in class_students if s['intern_extern'] == 'extern')
                
                # Berechne globale Extern-Quote
                total_extern = sum(1 for s in students if s['intern_extern'] == 'extern')
                global_extern_ratio = total_extern / len(students) if students else 0.5
                
                if student['intern_extern'] == 'extern':
                    new_ratio = (extern_in_class + 1) / (len(class_students) + 1)
                else:
                    new_ratio = extern_in_class / (len(class_students) + 1)
                
                extern_penalty = abs(global_extern_ratio - new_ratio) * 40
                score -= extern_penalty
            
            # Bisherige Klasse aufbrechen (nur für interne)
            if student['intern_extern'] == 'intern' and student['bisherige_klasse']:
                same_old_class_count = sum(
                    1 for s in class_students 
                    if s.get('bisherige_klasse') == student['bisherige_klasse']
                )
                score -= same_old_class_count * 10
            
<<<<<<< HEAD
            # Wunschpartner: Bonus, wenn gewünschte Freunde bereits in dieser Klasse sind
            wunsch_ids = student.get('wunsch_ids') or []
            if wunsch_ids:
                friends_in_class = sum(1 for s in class_students if s['id'] in wunsch_ids)
                score += friends_in_class * 60
                # Zusätzlicher Bonus, wenn der Wunsch beidseitig ist
                for s in class_students:
                    if s['id'] in wunsch_ids and student['id'] in (s.get('wunsch_ids') or []):
                        score += 20
=======
            # Reli/Ethik - Fokus beachten, sonst gleichmäßig verteilen
            reli_focus = config.get('reli_focus', '')
            student_reli = student.get('reli_ethik', '')
            
            if reli_focus:
                # Klasse hat Reli-Fokus
                if student_reli == reli_focus:
                    score += 25  # Bonus: Schüler passt zum Fokus
                else:
                    score -= 15  # Strafe: Schüler passt nicht
            elif class_students and student_reli:
                # Kein Fokus -> gleichmäßig verteilen
                reli_in_class = sum(1 for s in class_students if s.get('reli_ethik') == student_reli)
                total_reli = sum(1 for s in students if s.get('reli_ethik') == student_reli)
                global_reli_ratio = total_reli / len(students) if students else 0
                
                new_ratio = (reli_in_class + 1) / (len(class_students) + 1)
                reli_penalty = abs(global_reli_ratio - new_ratio) * 35
                score -= reli_penalty
            
            # Zusatzfach - Fokus beachten, sonst gleichmäßig verteilen
            zf_focus = config.get('zf_focus', '')
            student_zf = student.get('zusatzfach', '')
            
            if zf_focus:
                # Klasse hat ZF-Fokus
                if student_zf == zf_focus:
                    score += 25  # Bonus: Schüler passt zum Fokus
                else:
                    score -= 15  # Strafe: Schüler passt nicht
            elif class_students and student_zf:
                # Kein Fokus -> gleichmäßig verteilen
                zf_in_class = sum(1 for s in class_students if s.get('zusatzfach') == student_zf)
                total_zf = sum(1 for s in students if s.get('zusatzfach') == student_zf)
                global_zf_ratio = total_zf / len(students) if students else 0
                
                new_ratio = (zf_in_class + 1) / (len(class_students) + 1)
                zf_penalty = abs(global_zf_ratio - new_ratio) * 30
                score -= zf_penalty
>>>>>>> 12d6a2ed4dd06facda55ba5696c9a5b42f786793
            
            # Klassengröße - starke Strafe wenn über Zielgröße
            current_size = len(class_students)
            max_size = max_class_sizes[class_id]
            if current_size >= max_size:
                score -= 1000  # Klasse ist voll
            elif current_size > target_size:
                score -= (current_size - target_size) * 20
            
            return score
        
        def assign_student_to_best_class(student, eligible_classes):
            """Weist einen Schüler der besten Klasse aus eligible_classes zu"""
            best_class = None
            best_score = float('-inf')
            
            for class_id in eligible_classes:
                # Prüfe ob Klasse noch Platz hat
                if len(classes[class_id]) >= max_class_sizes[class_id]:
                    continue
                    
                score = assignment_score(student, class_id, classes)
                
                if score > best_score:
                    best_score = score
                    best_class = class_id
            
            # Fallback: Wenn alle Klassen "voll" sind, nimm die mit wenigsten Schülern
            if best_class is None and eligible_classes:
                best_class = min(eligible_classes, key=lambda c: len(classes[c]))
            
            if best_class is not None:
                student['assigned_class'] = best_class
                classes[best_class].append(student)
                return True
            return False
        
        # Schritt 2: Weise Kunst-Schüler zu
        for student in ku_students:
            # Erst Kunst-Klassen, dann gemischte Klassen
            eligible = ku_classes + mixed_classes if ku_classes else mixed_classes
            if not eligible:
                eligible = list(range(num_classes))  # Fallback: alle Klassen
            assign_student_to_best_class(student, eligible)
        
        # Schritt 3: Weise Musik-Schüler zu
        for student in mu_students:
            # Erst Musik-Klassen, dann gemischte Klassen
            eligible = mu_classes + mixed_classes if mu_classes else mixed_classes
            if not eligible:
                eligible = list(range(num_classes))  # Fallback: alle Klassen
            assign_student_to_best_class(student, eligible)
        
        # Schritt 4: Weise Schüler ohne Ku/Mu zu (alle Klassen möglich)
        for student in other_students:
            assign_student_to_best_class(student, list(range(num_classes)))
        
        class_assignment_data['classes'] = classes
        
        # Berechne Statistiken pro Klasse
        class_stats = []
        for class_id, class_students in classes.items():
            config = class_config[class_id]
            stats = {
                'id': class_id,
                'name': config.get('name', f'Klasse {class_id + 1}'),
                'language_focus': config.get('language_focus', ''),
                'art_music': config.get('art_music', ''),
                'total': len(class_students),
                'male': sum(1 for s in class_students if s['geschlecht'] == 'm'),
                'female': sum(1 for s in class_students if s['geschlecht'] == 'w'),
                'intern': sum(1 for s in class_students if s['intern_extern'] == 'intern'),
                'extern': sum(1 for s in class_students if s['intern_extern'] == 'extern'),
                'students': class_students
            }
            
            # Sprachstatistik
            lang_stats = defaultdict(int)
            for s in class_students:
                if s['fremdsprache']:
                    lang_stats[s['fremdsprache']] += 1
            stats['languages'] = dict(lang_stats)
            
            class_stats.append(stats)
        
        # Kurs-Zusammenlegungsanalyse
        MIN_COURSE_SIZE = 12
        
        course_merges = []
        
        def analyze_course_merges(category_key, category_values, label_map=None):
            """Analysiert ob Kurse klassenübergreifend zusammengelegt werden sollten.
            Bildet Gruppen aus beliebig vielen Klassen, bis Mindestgröße erreicht ist."""
            for val in category_values:
                classes_with_val = []
                for class_id, class_students_list in classes.items():
                    count = sum(1 for s in class_students_list if s.get(category_key) == val)
                    if count > 0:
                        cfg = class_config[class_id]
                        classes_with_val.append({
                            'class_id': class_id,
                            'class_name': cfg.get('name', f'Klasse {class_id + 1}'),
                            'count': count
                        })
                
                # Alle Klassen die alleine zu klein sind
                small = [c for c in classes_with_val if c['count'] < MIN_COURSE_SIZE]
                
                if len(small) < 2:
                    continue
                
                # Sortiere nach Größe absteigend (große zuerst paaren, effizienter)
                small.sort(key=lambda c: c['count'], reverse=True)
                
                display_name = label_map.get(val, val) if label_map else val
                used = set()
                
                # Greedy-Gruppierung: Nehme Klassen in Gruppen zusammen bis >= MIN_COURSE_SIZE
                while True:
                    remaining = [i for i, c in enumerate(small) if i not in used]
                    if len(remaining) < 2:
                        break
                    
                    # Starte neue Gruppe mit der größten verbleibenden Klasse
                    group = [remaining[0]]
                    group_count = small[remaining[0]]['count']
                    used.add(remaining[0])
                    
                    # Füge weitere Klassen hinzu bis Mindestgröße erreicht
                    for idx in remaining[1:]:
                        if group_count >= MIN_COURSE_SIZE:
                            break
                        group.append(idx)
                        group_count += small[idx]['count']
                        used.add(idx)
                    
                    # Nur Gruppen mit mindestens 2 Klassen ausgeben
                    if len(group) >= 2:
                        group_classes = [{'name': small[i]['class_name'], 'count': small[i]['count']} for i in group]
                        parts = [f'{small[i]["class_name"]} ({small[i]["count"]})' for i in group]
                        reason = ' + '.join(parts) + f' = {group_count} Schüler'
                        
                        course_merges.append({
                            'course': display_name,
                            'type': val,
                            'classes': group_classes,
                            'combined_count': group_count,
                            'reason': reason
                        })
        
        analyze_course_merges('kunst_musik', ['Ku', 'Mu'], {'Ku': 'Kunst', 'Mu': 'Musik'})
        analyze_course_merges('reli_ethik', ['Rev', 'Rka', 'Eth'], {'Rev': 'Evangelisch', 'Rka': 'Katholisch', 'Eth': 'Ethik'})
        analyze_course_merges('zusatzfach', ['Ek', 'Phil', 'Info', 'Spa', 'Bili'], {'Ek': 'Erdkunde', 'Phil': 'Philosophie', 'Info': 'Informatik', 'Spa': 'Spanisch', 'Bili': 'Bilingual'})
        
        return jsonify({
            'success': True,
            'classes': class_stats,
            'course_merges': course_merges,
            'min_course_size': MIN_COURSE_SIZE
        })
    
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/klassenzuteilung/move', methods=['POST'])
def move_student():
    """Verschiebt einen Schüler in eine andere Klasse"""
    global class_assignment_data
    
    try:
        data = request.get_json()
        student_id = data.get('studentId')
        target_class = data.get('targetClass')
        
        if student_id is None or target_class is None:
            return jsonify({'success': False, 'error': 'studentId und targetClass erforderlich'}), 400
        
        # Finde Schüler
        student = None
        for s in class_assignment_data['students']:
            if s['id'] == student_id:
                student = s
                break
        
        if not student:
            return jsonify({'success': False, 'error': 'Schüler nicht gefunden'}), 404
        
        old_class = student['assigned_class']
        
        # Entferne aus alter Klasse
        if old_class is not None and old_class in class_assignment_data['classes']:
            class_assignment_data['classes'][old_class] = [
                s for s in class_assignment_data['classes'][old_class] 
                if s['id'] != student_id
            ]
        
        # Füge zu neuer Klasse hinzu
        student['assigned_class'] = target_class
        if target_class not in class_assignment_data['classes']:
            class_assignment_data['classes'][target_class] = []
        class_assignment_data['classes'][target_class].append(student)
        
        # Berechne neue Statistiken
        class_stats = []
        for class_id, class_students in class_assignment_data['classes'].items():
            config = class_assignment_data['class_config'][class_id] if class_id < len(class_assignment_data['class_config']) else {}
            stats = {
                'id': class_id,
                'name': config.get('name', f'Klasse {class_id + 1}'),
                'language_focus': config.get('language_focus', ''),
                'art_music': config.get('art_music', ''),
                'total': len(class_students),
                'male': sum(1 for s in class_students if s['geschlecht'] == 'm'),
                'female': sum(1 for s in class_students if s['geschlecht'] == 'w'),
                'intern': sum(1 for s in class_students if s['intern_extern'] == 'intern'),
                'extern': sum(1 for s in class_students if s['intern_extern'] == 'extern'),
                'students': class_students
            }
            
            lang_stats = defaultdict(int)
            for s in class_students:
                if s['fremdsprache']:
                    lang_stats[s['fremdsprache']] += 1
            stats['languages'] = dict(lang_stats)
            
            class_stats.append(stats)
        
        return jsonify({
            'success': True,
            'classes': class_stats
        })
    
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/klassenzuteilung/export', methods=['GET'])
def export_class_assignment():
    """Exportiert die Klassenzuteilung als Excel"""
    global class_assignment_data
    
    try:
        students = class_assignment_data['students']
        class_config = class_assignment_data['class_config']
        
        if not students:
            return jsonify({'success': False, 'error': 'Keine Daten zum Exportieren'}), 400
        
        # Erstelle DataFrame
        export_data = []
        for student in students:
            class_id = student.get('assigned_class')
            class_name = ''
            if class_id is not None and class_id < len(class_config):
                class_name = class_config[class_id].get('name', f'Klasse {class_id + 1}')
            
            export_data.append({
                'Nachname': student['nachname'],
                'Vorname': student['vorname'],
                'Geschlecht': student['geschlecht'],
                'Intern/Extern': student['intern_extern'],
                '2. Fremdsprache': student['fremdsprache'],
                'Kunst/Musik': student['kunst_musik'],
                'Reli/Ethik': student.get('reli_ethik', ''),
                'Zusatzfach': student.get('zusatzfach', ''),
                'Bisherige Klasse': student.get('bisherige_klasse', ''),
                'Wunschpartner': ', '.join(student.get('wunsch_names', [])),
                'Neue Klasse': class_name
            })
        
        df = pd.DataFrame(export_data)
        
        # Erstelle Excel-Datei im Speicher
        excel_buffer = io.BytesIO()
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Klassenzuteilung')
            
            worksheet = writer.sheets['Klassenzuteilung']
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        excel_buffer.seek(0)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        return send_file(
            excel_buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'klassenzuteilung_{timestamp}.xlsx'
        )
    
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/klassenzuteilung/clear', methods=['POST'])
def clear_class_assignment():
    """Löscht alle Klassenzuteilungsdaten"""
    global class_assignment_data
    
    class_assignment_data = {
        'students': [],
        'classes': {},
        'class_config': [],
        'invalid_students': []
    }
    
    return jsonify({'success': True, 'message': 'Daten gelöscht'})

# ============================================================================
# Ende KLASSENZUTEILUNG API
# ============================================================================

@app.route('/upload_excel', methods=['POST'])
def upload_excel():
    """Handles Excel file upload, removes red-colored rows, and returns Vorname/Nachname data"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'Keine Datei hochgeladen'}), 400
        
        file = request.files['file']
        
        if file.filename == '':
            return jsonify({'success': False, 'error': 'Keine Datei ausgewählt'}), 400
        
        if not file.filename.endswith('.xlsx'):
            return jsonify({'success': False, 'error': 'Nur .xlsx Dateien werden unterstützt'}), 400
        
        # Load the workbook
        wb = load_workbook(file, data_only=False)
        ws = wb.active
        
        # First row is the header - search for "Vorname" and "Nachname"
        vorname_col = None
        nachname_col = None
        
        first_row = list(ws[1])
        for col_idx, cell in enumerate(first_row, start=1):
            if cell.value:
                cell_value = str(cell.value).strip().lower()
                if 'vorname' in cell_value:
                    vorname_col = col_idx
                elif 'nachname' in cell_value:
                    nachname_col = col_idx
        
        if not vorname_col or not nachname_col:
            return jsonify({'success': False, 'error': 'Spalten "Vorname" und/oder "Nachname" nicht in der ersten Zeile gefunden'}), 400
        
        # Process rows starting from row 2 and filter out red-colored rows
        data = []
        for row in ws.iter_rows(min_row=2):
            # Check if any cell in the row is colored red
            is_red = False
            for cell in row:
                if cell.fill and cell.fill.start_color:
                    color = cell.fill.start_color
                    if hasattr(color, 'rgb') and color.rgb:
                        rgb_str = str(color.rgb)
                        # Remove alpha channel if present (ARGB format starts with FF)
                        if len(rgb_str) == 8:
                            rgb_str = rgb_str[2:]
                        
                        try:
                            r = int(rgb_str[0:2], 16)
                            g = int(rgb_str[2:4], 16)
                            b = int(rgb_str[4:6], 16)
                            # Consider it red if red component is dominant
                            if r > 180 and r > (g + 50) and r > (b + 50):
                                is_red = True
                                break
                        except (ValueError, IndexError):
                            pass
            
            if not is_red:
                # Extract Vorname and Nachname
                vorname = row[vorname_col - 1].value if len(row) >= vorname_col else None
                nachname = row[nachname_col - 1].value if len(row) >= nachname_col else None
                
                # Only add if at least one field has a value
                if vorname or nachname:
                    data.append({
                        'Vorname': str(vorname).strip() if vorname else '',
                        'Nachname': str(nachname).strip() if nachname else ''
                    })
        
        return jsonify({
            'success': True,
            'data': data,
            'rowCount': len(data)
        })
    
    except Exception as e:
        print(f"Error processing Excel file: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': f'Fehler beim Verarbeiten der Datei: {str(e)}'}), 500

def is_port_available(port):
    """Prüft ob ein Port verfügbar ist"""
    sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    sock.settimeout(1)
    try:
        sock.bind(('0.0.0.0', port))
        sock.close()
        return True
    except (socket.error, OSError):
        return False

def find_available_port(min_port=5000, max_port=5999, max_attempts=50):
    """Findet einen verfügbaren zufälligen Port im angegebenen Bereich"""
    attempts = 0
    tried_ports = set()
    
    while attempts < max_attempts:
        port = random.randint(min_port, max_port)
        
        # Überspringe bereits getestete Ports
        if port in tried_ports:
            continue
        
        tried_ports.add(port)
        attempts += 1
        
        if is_port_available(port):
            return port
    
    # Fallback: Durchsuche alle Ports sequentiell
    for port in range(min_port, max_port + 1):
        if port not in tried_ports and is_port_available(port):
            return port
    
    # Wenn kein Port gefunden wurde, verwende einen vom System zugewiesenen Port
    return None

if __name__ == '__main__':
    print("="*80)
    print("🚀 Methodentag Kurszuteilung - Web Interface")
    print("="*80)
    print(f"\n🖥️  Betriebssystem: {platform.system()}")
    
    if IS_RAM_ONLY_MODE:
        print(f"\n🔒 DATENSCHUTZMODUS (RAM-Only):")
        print(f"   • Keine persistente Speicherung")
        print(f"   • Alle Daten nur im RAM")
        print(f"   • Dateien werden nach Server-Neustart gelöscht")
    else:
        print(f"\n💾 Datenspeicherung:")
        print(f"   • Uploads: {UPLOADS_DIR}")
        print(f"   • Ergebnisse: {RESULTS_DIR}")
        print(f"   • Nicht über Web-Interface zugänglich")
    
    # Finde verfügbaren Port
    port = find_available_port()
    
    if port is None:
        print("\n⚠️  Warnung: Kein freier Port im Bereich 5000-5999 gefunden!")
        print("   Verwende Port 0 (System wählt automatisch)")
        port = 0
    
    url = f"http://localhost:{port}"
    print(f"\n🌐 Server läuft auf: {url}")
    print("\n" + "="*80 + "\n")
    
    # Öffne Browser automatisch nach kurzer Verzögerung
    def open_browser():
        import time
        time.sleep(1.5)  # Warte bis Server bereit ist
        try:
            webbrowser.open(url)
            print(f"🌐 Browser geöffnet: {url}")
        except Exception as e:
            print(f"⚠️  Konnte Browser nicht öffnen: {e}")
    
    threading.Thread(target=open_browser, daemon=True).start()
    
    app.run(debug=True, host='0.0.0.0', port=port, use_reloader=False)
